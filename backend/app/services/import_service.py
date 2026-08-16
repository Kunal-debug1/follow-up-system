from __future__ import annotations

import csv
import re
from itertools import islice
from pathlib import Path
from time import monotonic
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..models import Customer, ImportBatch


# ============================================================
# IMPORT CONFIGURATION
# ============================================================
# These limits protect the free Render instance from very large
# Excel/CSV files consuming too much memory or CPU.
MAX_WORKSHEET_ROWS = 50_000
MAX_COLUMNS = 80
MAX_HEADER_SCAN_ROWS = 30
MAX_SAMPLE_ROWS = 100
IMPORT_BATCH_SIZE = 250


# ============================================================
# EXCEL COLUMN ALIASES
# ============================================================
# The importer automatically detects common column names.
#
# Example:
# "Mobile Number", "Phone", "Contact Number"
# will all be mapped to the "phone" field.
FIELD_ALIASES = {
    "name": (
        "name",
        "customer name",
        "consumer name",
        "client name",
        "full name",
    ),
    "phone": (
        "phone",
        "mobile",
        "mobile number",
        "contact",
        "contact no",
        "contact number",
        "phone number",
        "telephone",
    ),
    "email": (
        "email",
        "email address",
        "e mail",
    ),
    "consumer_number": (
        "consumer number",
        "consumer no",
        "consumer id",
        "account number",
        "account no",
    ),
    "service": (
        "service",
        "product",
        "service name",
        "requirement",
        "trf desc",
    ),
    "region": (
        "region",
        "region name",
    ),
    "zone": (
        "zone",
        "zone name",
    ),
    "circle": (
        "circle",
        "circle name",
    ),
    "division": (
        "division",
        "division name",
    ),
    "subdivision": (
        "subdivision",
        "sub division",
        "subdivision name",
    ),
    "business_unit": (
        "bu",
        "business unit",
    ),
}


# Fields that are actually stored on Customer.
CUSTOMER_FIELDS = (
    "name",
    "phone",
    "email",
    "service",
    "consumer_number",
    "address",
    "region",
    "zone",
    "circle",
    "division",
    "subdivision",
    "business_unit",
)


# ============================================================
# BASIC NORMALIZATION HELPERS
# ============================================================

def normalize_column(value: Any) -> str:
    """
    Convert a column heading into a predictable format.

    Example:
        "Customer  Name" -> "customer name"
        "Mobile_Number"  -> "mobile number"
    """
    return re.sub(
        r"\s+",
        " ",
        re.sub(
            r"[^a-z0-9]+",
            " ",
            str(value).strip().lower(),
        ),
    ).strip()


def clean_value(value: Any) -> str | None:
    """
    Convert Excel/CSV values into clean strings.

    Empty values such as:
        None
        NaN
        null
        <NA>
        000000000
    are treated as empty.
    """
    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    if text.lower() in {
        "nan",
        "none",
        "null",
        "nat",
        "<na>",
    }:
        return None

    # Excel can convert numeric values such as 12345 into "12345.0".
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]

    # Ignore obviously invalid zero-only values.
    if re.fullmatch(r"0{8,}", re.sub(r"\s+", "", text)):
        return None

    return text


def normalize_phone(value: str | None) -> str | None:
    """
    Normalize Indian mobile numbers to a 10-digit format.

    Supported examples:
        9876543210
        09876543210
        919876543210
        +91 9876543210
    """
    if not value:
        return None

    digits = re.sub(r"\D", "", value)

    # Remove India country code.
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]

    # Remove leading zero.
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]

    # Accept only valid Indian 10-digit mobile numbers.
    if len(digits) == 10 and digits[0] in "6789":
        return digits

    return None


# ============================================================
# SAFETY LIMIT FOR ROWS/COLUMNS
# ============================================================

def _bounded_rows(
    rows: Iterable[tuple[Any, ...]],
) -> Iterator[tuple[int, tuple[Any, ...]]]:
    """
    Yield worksheet rows while enforcing safety limits.

    This prevents an accidentally huge Excel file from
    consuming excessive Render resources.
    """
    for number, row in enumerate(rows, start=1):

        if number > MAX_WORKSHEET_ROWS:
            raise ValueError(
                f"Worksheet exceeds the "
                f"{MAX_WORKSHEET_ROWS:,}-row safety limit."
            )

        # Only process the first MAX_COLUMNS columns.
        yield number, tuple(row[:MAX_COLUMNS])


# ============================================================
# XLSX WORKBOOK HELPERS
# ============================================================

def workbook_sheets(path: Path) -> list[str]:
    """
    Return all worksheet names from an XLSX file.

    IMPORTANT:
    openpyxl Workbook does not support:
        with load_workbook(...) as workbook

    Therefore we explicitly close the workbook in finally.
    This fixes the Render 500 error you encountered.
    """
    workbook = None

    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

        return list(workbook.sheetnames)

    except Exception as exc:
        raise ValueError(
            "Could not read XLSX workbook."
        ) from exc

    finally:
        # Always release the XLSX file resources.
        if workbook is not None:
            workbook.close()


def _sheet_rows(
    path: Path,
    sheet_name: str,
) -> Iterator[tuple[int, tuple[Any, ...]]]:
    """
    Read rows from one XLSX worksheet.

    The workbook is opened in read-only mode so large files
    don't need to be fully loaded into RAM.
    """
    workbook = None

    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                "Selected sheet does not exist."
            )

        worksheet = workbook[sheet_name]

        # Yield rows while enforcing our safety limits.
        yield from _bounded_rows(
            worksheet.iter_rows(values_only=True)
        )

    except ValueError:
        raise

    except Exception as exc:
        raise ValueError(
            "Could not parse XLSX worksheet."
        ) from exc

    finally:
        # Critical for Render memory/resource management.
        if workbook is not None:
            workbook.close()


# ============================================================
# CSV READER
# ============================================================

def _csv_rows(
    path: Path,
) -> Iterator[tuple[int, tuple[Any, ...]]]:
    """
    Read UTF-8 CSV files safely.
    """
    try:
        with path.open(
            "r",
            encoding="utf-8-sig",
            newline="",
        ) as handle:

            yield from _bounded_rows(
                csv.reader(handle)
            )

    except UnicodeDecodeError as exc:
        raise ValueError(
            "CSV must be UTF-8 encoded."
        ) from exc

    except csv.Error as exc:
        raise ValueError(
            "Could not parse CSV file."
        ) from exc


# ============================================================
# HEADER DETECTION
# ============================================================

def _header_score(
    row: tuple[Any, ...],
) -> int:
    """
    Score a row to determine whether it is probably
    the Excel header row.
    """
    text = " | ".join(
        normalize_column(value)
        for value in row
        if clean_value(value)
    )

    keywords = (
        "consumer number",
        "consumer name",
        "customer name",
        "contact",
        "phone",
        "mobile",
        "email",
        "address",
    )

    score = sum(
        2
        for item in keywords
        if item in text
    )

    # Strong indicators for the user's existing data.
    if "consumer number" in text:
        score += 5

    if "consumer name" in text:
        score += 5

    return score


def _headers_and_mapping(
    scanned: list[tuple[int, tuple[Any, ...]]],
) -> tuple[int, list[str], dict[str, Any]]:
    """
    Detect the header row and map Excel columns to CRM fields.
    """
    if not scanned:
        raise ValueError(
            "The selected sheet is empty."
        )

    # Pick the row with the highest header score.
    header_number, header_values = max(
        scanned[:MAX_HEADER_SCAN_ROWS],
        key=lambda item: _header_score(item[1]),
    )

    if not any(
        clean_value(value)
        for value in header_values
    ):
        raise ValueError(
            "The selected sheet is empty."
        )

    # Give unnamed columns a predictable name.
    headers = [
        clean_value(value)
        or f"Unnamed: {index}"
        for index, value in enumerate(header_values)
    ]

    normalized = {
        header: normalize_column(header)
        for header in headers
    }

    mapping: dict[str, Any] = {}

    # --------------------------------------------------------
    # Generic aliases
    # --------------------------------------------------------
    for field, aliases in FIELD_ALIASES.items():

        aliases_normalized = {
            normalize_column(alias)
            for alias in aliases
        }

        for header, normalized_value in normalized.items():

            if normalized_value in aliases_normalized:
                mapping[field] = header
                break

    # --------------------------------------------------------
    # Exact headings used by the existing customer workbook.
    # --------------------------------------------------------
    exact_sources = {
        "name": "consumer_name",
        "consumer_number": "consumer_number",
        "email": "email_id",
        "service": "trf_desc",
        "region": "region_name",
        "zone": "zone_name",
        "circle": "circle_name",
        "division": "division_name",
        "subdivision": "subdivision_name",
        "business_unit": "bu",
    }

    for field, source in exact_sources.items():

        source_normalized = normalize_column(source)

        for header, value in normalized.items():

            if value == source_normalized:
                mapping[field] = header
                break

    # --------------------------------------------------------
    # Address can be spread across multiple Excel columns.
    # --------------------------------------------------------
    address_columns = [
        header
        for header, value in normalized.items()
        if value in {
            "address",
            "address 1",
            "address 2",
            "address 3",
            "address l1",
            "address l2",
            "address l3",
        }
    ]

    if address_columns:
        mapping["address_columns"] = address_columns

    return (
        header_number,
        headers,
        mapping,
    )


# ============================================================
# SCAN EXCEL/CSV SOURCE
# ============================================================

def _scan_source(
    path: Path,
    file_type: str,
    sheet_name: str | None,
) -> tuple[
    list[tuple[int, tuple[Any, ...]]],
    str | None,
]:
    """
    Read only enough rows to detect headers and generate
    a small preview.
    """
    if file_type == "xlsx":

        if sheet_name is None:
            sheets = workbook_sheets(path)

            sheet_name = (
                sheets[0]
                if sheets
                else None
            )

        if not sheet_name:
            raise ValueError(
                "Workbook contains no worksheets."
            )

        source = _sheet_rows(
            path,
            sheet_name,
        )

    else:
        source = _csv_rows(path)

    rows = []

    for row in source:

        rows.append(row)

        if len(rows) >= (
            MAX_HEADER_SCAN_ROWS
            + MAX_SAMPLE_ROWS
        ):
            break

    return rows, sheet_name


# ============================================================
# FILE ANALYSIS
# ============================================================

def analyze_file(
    path: Path,
    file_type: str,
    sheet_name: str | None = None,
) -> dict:
    """
    Analyze an uploaded Excel/CSV file without importing it.
    """
    scanned, selected_sheet = _scan_source(
        path,
        file_type,
        sheet_name,
    )

    header_row, headers, mapping = (
        _headers_and_mapping(scanned)
    )

    # Count rows after detected header.
    header_index = next(
        (
            i
            for i, item in enumerate(scanned)
            if item[0] == header_row
        ),
        len(scanned),
    )

    sample_rows = max(
        0,
        len(scanned)
        - header_index
        - 1,
    )

    return {
        "total_rows": sample_rows,
        "columns": headers,
        "header_row": header_row,
        "detected_mapping": mapping,
        "missing_required": (
            []
            if "name" in mapping
            else ["name"]
        ),
        "selected_sheet": selected_sheet,
    }


# ============================================================
# CONVERT SOURCE ROWS INTO CRM RECORDS
# ============================================================

def iter_records(
    path: Path,
    file_type: str,
    sheet_name: str | None,
    mapping: dict[str, Any],
) -> Iterator[dict]:
    """
    Convert Excel/CSV rows into normalized CRM customer records.
    """
    if file_type == "xlsx":

        if not sheet_name:
            raise ValueError(
                "A worksheet must be selected."
            )

        source = _sheet_rows(
            path,
            sheet_name,
        )

    else:
        source = _csv_rows(path)

    rows = iter(source)

    # Read only enough rows to locate the header.
    initial = list(
        islice(
            rows,
            MAX_HEADER_SCAN_ROWS,
        )
    )

    if not initial:
        return

    header_number, headers, _ = (
        _headers_and_mapping(initial)
    )

    # Find where data starts.
    start_index = next(
        i
        for i, item in enumerate(initial)
        if item[0] == header_number
    ) + 1

    row_stream = iter(
        initial[start_index:]
    )

    def remaining():
        # First return already-read rows.
        yield from row_stream

        # Then continue reading the source.
        yield from rows

    for number, values in remaining():

        # Convert row into:
        # {
        #   "Excel Column": "value"
        # }
        row = {
            headers[index]:
                values[index]
                if index < len(values)
                else None
            for index in range(len(headers))
        }

        name = clean_value(
            row.get(
                mapping.get("name", "")
            )
        )

        phone = clean_value(
            row.get(
                mapping.get("phone", "")
            )
        )

        # Ignore completely empty records.
        if not name and not phone:
            continue

        # Combine multiple address columns.
        address_parts = []

        for column in mapping.get(
            "address_columns",
            [],
        ):
            value = clean_value(
                row.get(column)
            )

            if value:
                address_parts.append(value)

        address = (
            ", ".join(
                dict.fromkeys(address_parts)
            )
            if address_parts
            else None
        )

        yield {
            "name": name or "Unknown",

            "phone": normalize_phone(
                phone
            ),

            "email": clean_value(
                row.get(
                    mapping.get("email", "")
                )
            ),

            "service": clean_value(
                row.get(
                    mapping.get("service", "")
                )
            ),

            "consumer_number": clean_value(
                row.get(
                    mapping.get(
                        "consumer_number",
                        "",
                    )
                )
            ),

            "address": address,

            "region": clean_value(
                row.get(
                    mapping.get(
                        "region",
                        "",
                    )
                )
            ),

            "zone": clean_value(
                row.get(
                    mapping.get(
                        "zone",
                        "",
                    )
                )
            ),

            "circle": clean_value(
                row.get(
                    mapping.get(
                        "circle",
                        "",
                    )
                )
            ),

            "division": clean_value(
                row.get(
                    mapping.get(
                        "division",
                        "",
                    )
                )
            ),

            "subdivision": clean_value(
                row.get(
                    mapping.get(
                        "subdivision",
                        "",
                    )
                )
            ),

            "business_unit": clean_value(
                row.get(
                    mapping.get(
                        "business_unit",
                        "",
                    )
                )
            ),

            # Keep original Excel row number.
            "source_row": number,
        }


# ============================================================
# DATABASE LOOKUPS
# ============================================================

def _load_by_consumer(
    db: Session,
    values: set[str],
) -> dict[str, Customer]:
    """
    Find existing customers by consumer number.
    """
    if not values:
        return {}

    customers = (
        db.query(Customer)
        .filter(
            Customer.consumer_number.in_(values)
        )
        .order_by(Customer.id)
        .all()
    )

    return {
        customer.consumer_number: customer
        for customer in customers
        if customer.consumer_number
    }


def _load_by_phone_name(
    db: Session,
    phones: set[str],
) -> dict[tuple[str, str], Customer]:
    """
    Fallback duplicate detection using:
        phone + normalized customer name
    """
    if not phones:
        return {}

    customers = (
        db.query(Customer)
        .filter(
            Customer.phone.in_(phones)
        )
        .order_by(Customer.id)
        .all()
    )

    return {
        (
            customer.phone,
            normalize_column(customer.name),
        ): customer
        for customer in customers
        if customer.phone
    }


# ============================================================
# CUSTOMER UPDATE DETECTION
# ============================================================

def _has_updates(
    customer: Customer,
    record: dict,
) -> bool:
    """
    Check whether an imported record contains
    information that differs from the database.
    """
    return any(
        record.get(field)
        and getattr(customer, field)
        != record[field]
        for field in CUSTOMER_FIELDS
    )


# ============================================================
# SIMPLE PREVIEW
# ============================================================

def preview_records(
    db: Session,
    records: Iterable[dict],
    limit: int,
) -> tuple[dict, list[dict]]:
    """
    Preview a prepared collection of records.
    """
    collected = list(records)

    consumers = {
        record["consumer_number"]
        for record in collected
        if record.get("consumer_number")
    }

    existing = _load_by_consumer(
        db,
        consumers,
    )

    seen: set[str] = set()

    duplicates = 0
    existing_rows = 0
    updates = 0

    for record in collected:

        consumer = record.get(
            "consumer_number"
        )

        if not consumer:
            continue

        if consumer in seen:
            duplicates += 1
            continue

        seen.add(consumer)

        if consumer in existing:
            existing_rows += 1

            updates += int(
                _has_updates(
                    existing[consumer],
                    record,
                )
            )

    return (
        {
            "duplicate_rows_in_file": duplicates,
            "already_in_database": existing_rows,
            "update_rows": updates,
            "new_records": max(
                0,
                len(collected)
                - duplicates
                - existing_rows,
            ),
        },
        collected[:limit],
    )


# ============================================================
# FULL FILE PREVIEW
# ============================================================

def preview_file(
    db: Session,
    path: Path,
    file_type: str,
    sheet_name: str | None,
    mapping: dict[str, Any],
    limit: int,
) -> tuple[dict, list[dict]]:
    """
    Preview the file without importing it.

    The file is read in bounded passes rather than loading
    the entire Excel file into memory.
    """

    consumers: set[str] = set()
    fallback_phones: set[str] = set()

    total = 0
    with_phone = 0
    duplicates = 0

    seen_consumers: set[str] = set()
    seen_fallback: set[
        tuple[str, str]
    ] = set()

    preview: list[dict] = []

    # --------------------------------------------------------
    # First pass:
    # Count rows and detect duplicates.
    # --------------------------------------------------------
    for record in iter_records(
        path,
        file_type,
        sheet_name,
        mapping,
    ):

        total += 1

        if record.get("phone"):
            with_phone += 1

        consumer = record.get(
            "consumer_number"
        )

        if consumer:

            if consumer in seen_consumers:
                duplicates += 1

            seen_consumers.add(consumer)
            consumers.add(consumer)

        elif record.get("phone"):

            key = (
                record["phone"],
                normalize_column(
                    record["name"]
                ),
            )

            if key in seen_fallback:
                duplicates += 1

            seen_fallback.add(key)
            fallback_phones.add(
                record["phone"]
            )

        # Keep only the requested preview rows.
        if len(preview) < limit:
            preview.append(record)

    # --------------------------------------------------------
    # Load matching database records.
    # --------------------------------------------------------
    by_consumer = _load_by_consumer(
        db,
        consumers,
    )

    by_fallback = _load_by_phone_name(
        db,
        fallback_phones,
    )

    existing_rows = 0
    updates = 0

    counted_consumers: set[str] = set()
    counted_fallback: set[
        tuple[str, str]
    ] = set()

    # --------------------------------------------------------
    # Second pass:
    # Determine existing/update records.
    # --------------------------------------------------------
    for record in iter_records(
        path,
        file_type,
        sheet_name,
        mapping,
    ):

        consumer = record.get(
            "consumer_number"
        )

        if consumer:

            if consumer in counted_consumers:
                continue

            counted_consumers.add(consumer)

            existing = by_consumer.get(
                consumer
            )

        elif record.get("phone"):

            key = (
                record["phone"],
                normalize_column(
                    record["name"]
                ),
            )

            if key in counted_fallback:
                continue

            counted_fallback.add(key)

            existing = by_fallback.get(key)

        else:
            continue

        if existing:
            existing_rows += 1

            updates += int(
                _has_updates(
                    existing,
                    record,
                )
            )

    return (
        {
            "rows_in_file": total,
            "valid_records": total,
            "records_with_phone": with_phone,
            "records_without_phone":
                total - with_phone,
            "duplicate_rows_in_file": duplicates,
            "already_in_database": existing_rows,
            "update_rows": updates,
            "new_records": max(
                0,
                total
                - duplicates
                - existing_rows,
            ),
        },
        preview,
    )


# ============================================================
# IMPORT RECORDS IN BATCHES
# ============================================================

def import_record_batches(
    db: Session,
    filename: str,
    file_type: str,
    records: Iterable[dict],
) -> dict:
    """
    Import customers into PostgreSQL in small batches.

    Batching keeps memory usage low and works better
    with Render's free instance.
    """

    started = monotonic()

    # Create an import history record.
    batch = ImportBatch(
        filename=filename,
        file_type=file_type,
        total_rows=0,
        status="processing",
    )

    db.add(batch)
    db.commit()
    db.refresh(batch)

    imported = 0
    updated = 0
    duplicates = 0
    skipped = 0
    total = 0

    seen_consumers: set[str] = set()

    seen_fallback: set[
        tuple[str, str]
    ] = set()

    try:

        pending: list[dict] = []

        # Collect only IMPORT_BATCH_SIZE records.
        for record in records:

            pending.append(record)

            if len(pending) >= IMPORT_BATCH_SIZE:

                counts = _import_batch(
                    db,
                    batch,
                    pending,
                    seen_consumers,
                    seen_fallback,
                )

                total += len(pending)
                imported += counts[0]
                updated += counts[1]
                duplicates += counts[2]
                skipped += counts[3]

                pending = []

        # Import remaining records.
        if pending:

            counts = _import_batch(
                db,
                batch,
                pending,
                seen_consumers,
                seen_fallback,
            )

            total += len(pending)
            imported += counts[0]
            updated += counts[1]
            duplicates += counts[2]
            skipped += counts[3]

        # Save final import statistics.
        batch.total_rows = total
        batch.imported_rows = imported
        batch.duplicate_rows = duplicates
        batch.skipped_rows = skipped
        batch.status = "completed"

        db.commit()

    except Exception:

        # Roll back failed transaction.
        db.rollback()

        batch.status = "failed"

        db.add(batch)
        db.commit()

        raise

    return {
        "import_id": batch.id,
        "total_rows": total,
        "imported_rows": imported,
        "updated_rows": updated,
        "duplicate_rows": duplicates,
        "skipped_rows": skipped,
        "status": "completed",
        "processing_seconds": round(
            monotonic() - started,
            3,
        ),
    }


# ============================================================
# IMPORT ONE DATABASE BATCH
# ============================================================

def _import_batch(
    db: Session,
    batch: ImportBatch,
    records: list[dict],
    seen_consumers: set[str],
    seen_fallback: set[tuple[str, str]],
) -> tuple[int, int, int, int]:
    """
    Import one small batch of records.

    Returns:
        imported
        updated
        duplicates
        skipped
    """

    consumers = {
        record["consumer_number"]
        for record in records
        if record.get("consumer_number")
    }

    phones = {
        record["phone"]
        for record in records
        if (
            not record.get(
                "consumer_number"
            )
            and record.get("phone")
        )
    }

    # Fetch existing customers once per batch.
    by_consumer = _load_by_consumer(
        db,
        consumers,
    )

    by_fallback = _load_by_phone_name(
        db,
        phones,
    )

    imported = 0
    updated = 0
    duplicates = 0
    skipped = 0

    try:

        for record in records:

            consumer = record.get(
                "consumer_number"
            )

            key = (
                record.get("phone") or "",
                normalize_column(
                    record["name"]
                ),
            )

            # ------------------------------------------------
            # Duplicate inside current import file.
            # ------------------------------------------------
            if (
                consumer
                and consumer in seen_consumers
            ):
                duplicates += 1
                continue

            if (
                not consumer
                and record.get("phone")
                and key in seen_fallback
            ):
                duplicates += 1
                continue

            # ------------------------------------------------
            # Find existing database customer.
            # ------------------------------------------------
            if consumer:
                existing = by_consumer.get(
                    consumer
                )

            elif record.get("phone"):
                existing = by_fallback.get(
                    key
                )

            else:
                existing = None

            # Mark record as processed.
            if consumer:
                seen_consumers.add(consumer)

            elif record.get("phone"):
                seen_fallback.add(key)

            # ------------------------------------------------
            # Existing customer:
            # update only fields containing new information.
            # ------------------------------------------------
            if existing:

                changed = 0

                for field in CUSTOMER_FIELDS:

                    value = record.get(field)

                    if (
                        value
                        and getattr(
                            existing,
                            field,
                        ) != value
                    ):
                        setattr(
                            existing,
                            field,
                            value,
                        )

                        changed += 1

                if changed:
                    updated += 1
                else:
                    skipped += 1

                continue

            # ------------------------------------------------
            # New customer.
            # ------------------------------------------------
            customer_data = {
                field: record.get(field)
                for field in CUSTOMER_FIELDS
            }

            customer = Customer(
                **customer_data,
                import_id=batch.id,
                source_file=batch.filename,
                source_row=record["source_row"],
            )

            db.add(customer)

            imported += 1

        # Commit this batch.
        db.commit()

    except Exception:

        # Roll back this batch if anything fails.
        db.rollback()

        raise

    return (
        imported,
        updated,
        duplicates,
        skipped,
    )
